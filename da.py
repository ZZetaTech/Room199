#import pandas as pd

#df = pd.read_excel(r'C:/Users/lsang/Desktop/ZetaTechSheet.xlsx')
import openpyxl

# LPP: this path ain't gonna work; first of all, this path doesn't exist in our current filesystem; second of all, the path is written Windows-style (backslashes, etc.) instead of Linux-style (forward slashes)
path = r'C:\Users\lsang\Desktop\ZetaTechSheet.xlsx'
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row=1, column=1)
print(cell_obj.value)
#from openpyxl import Workbook
#from pathlib import Path
#wb = Workbook()
#filepath = "'C:/Users/lsang/Desktop/ZetaTechSheet.xlsx'"
#wb = openpyxl.Workbook()

#wb.save(filepath)
#ws =  wb.active
#ws.title = "Changed Sheet"
#xlsx_file = openpyxl.load_workbook(r"C:/Users/lsang/Desktop/ZetaTechSheet.xlsx")
#wb.save(filename = 'sample_book.xlsx')

#wb_obj = openpyxl.load_workbook(xlsx_file)

# Read the active sheet:
#sheet = wb_obj.active
#print(sheet["A1"].value)
